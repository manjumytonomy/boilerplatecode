
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def get_html_file_info(html_dir):
    html_file_info = []
    for root, dirs, files in os.walk(html_dir):
        for file in files:
            if file.endswith('.html'):
                relative_path = os.path.relpath(root, html_dir)
                html_file_info.append({
                    'Relative Path': relative_path,
                    'HTML File Name': file
                })
    return html_file_info

def get_pdf_file_info(pdf_dir):
    pdf_file_info = {}
    for root, dirs, files in os.walk(pdf_dir):
        for file in files:
            if file.endswith('.pdf'):
                # Remove the 'PDFs' directory from the relative path
                relative_path_components = root.split(os.sep)
                relative_path_components = [component for component in relative_path_components if component != 'PDFs']
                relative_path = os.path.join(*relative_path_components)
                pdf_file_info[relative_path] = file
    return pdf_file_info

def main():
    #html_dir = input("Enter the path to the HTML files directory: ")
    #pdf_dir = input("Enter the path to the PDF files directory: ")

    #html_dir='/Users/manju/Mytonomy Inc Dropbox/Manju Komarlu/Written Education - Upwork Share/WE-Tester01/AAH/Jan-2025-First100 Test-ENGLISH/HTML'
    #pdf_dir='/Users/manju/Work/WrittenEducation Project/Customer Releases/AAH/PDFs/Jan-2024-First100 Test-ENGLISH'

    html_dir = '/Users/manju/Work/WrittenEducation Project/Customer Releases/AAH/HTML/Jan-20-2025-AAH_Test batch-113 Files/HTML/Clara Project/ENGLISH'
    pdf_dir = '/Users/manju/Work/WrittenEducation Project/Customer Releases/AAH/PDFs/Jan-20-2025-AAH_Test batch-113 Files/ENGLISH'

    html_file_info = get_html_file_info(html_dir)
    pdf_file_info = get_pdf_file_info(pdf_dir)

    df = pd.DataFrame(html_file_info)
    df['PDF File Name'] = df['Relative Path'].apply(lambda x: pdf_file_info.get(x, 'Not Found'))

    with pd.ExcelWriter('List of 113 Files - Jan 20-2025.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        sheet = writer.sheets['Sheet1']
        for i in range(2, sheet.max_row + 1):
            if i % 2 == 0:
                sheet.cell(row=i, column=1).fill = PatternFill(start_color='FFC7C7', end_color='FFC7C7', fill_type='solid')
                sheet.cell(row=i, column=2).fill = PatternFill(start_color='FFC7C7', end_color='FFC7C7', fill_type='solid')
                sheet.cell(row=i, column=3).fill = PatternFill(start_color='FFC7C7', end_color='FFC7C7', fill_type='solid')
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

if __name__ == "__main__":
    main()

