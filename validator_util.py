import os 
import configparser
from bs4 import BeautifulSoup  #this library is used to parse the HTML 
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from langdetect import detect  # need to add this to the dependencies library 
import datetime
import logging
import re
import streamlit as st

script_dir = os.path.dirname(os.path.abspath(__file__))     # Get the absolute path of the directory that this file is in (i.e. Streamlit UI)
project_dir = os.path.dirname(script_dir)                   # Go up one directory

# Configuration
config = configparser.ConfigParser()
config.read(os.path.join(project_dir, 'config.ini'))
customer = config['CUSTOMER']['customer_name']

configCustomer = configparser.ConfigParser()
configCustomer.read(os.path.join(project_dir, f'CustomerConfigs/{customer}_config.ini'))

if not os.path.join(configCustomer[f'{customer}']['local_folder_path']):
    st.error(f'Invalid local_folder_path in customer config for customer {customer}')
    st.stop()

# The below 2 DIR path variables are assumed to be static irrespective of the active customer
FOLDER_STORAGE_DIR = os.path.join(configCustomer[f'{customer}']['local_folder_path'], "Folder Storage")
if not os.path.isdir(FOLDER_STORAGE_DIR):
    st.error(f'Please ensure that local_folder_path is valid and local_folder_path contains the folder "Folder Storage", and re-run the app.')
    st.stop()

EPIC_HTML_REQUIREMENTS_DIR = os.path.join(configCustomer[f'{customer}']['local_folder_path'], "EpicHtmlRequirements")
if not os.path.isdir(EPIC_HTML_REQUIREMENTS_DIR):
    st.error(f'Please ensure that local_folder_path is valid and local_folder_path contains the folder "EpicHtmlRequirements", and re-run the app.')
    st.stop()

# Mapping for standarizing column names, even if column names change in the excel sheet
COLS = {"customer": "Customer", 
        "logo": "Logo", 
        "condition_area": "Condition Area", 
        "filepath": "Filepath", 
        "title": "Title", 
        "unique_name": "Unique Name", 
        "keyword": "Keyword", 
        "diagnosis_code": "Diagnosis Code", 
        "cpt_code": "CPT Code", 
        "language": "Language", 
        "corresponding_language": "Corresponding Language", 
        "language_index": "Language Index", 
        "source": "Source", 
        "document_type": "Document Type", 
        "qr_code": "QR Code", 
        "short_url": "Short URL", 
        "customer_disclaimer": "Customer Disclaimer"}

def init_logging(log_config, dir_name):
    # --------------------------------GLOBAL LOGGING SETUP--------------------------------
    log_file = os.path.join(project_dir, config['LOGGING']['log_file'])

    # Get the root logger -- this part makes sure that all logs are directed to the same file 
    root_logger = logging.getLogger()

    # Clear existing handlers if they exist
    if root_logger.hasHandlers():
        root_logger.handlers.clear()

    # Set up the file handler with rotation
    handler = logging.handlers.RotatingFileHandler(log_file, maxBytes=int(config['LOGGING']['max_bytes']), backupCount=int(config['LOGGING']['backup_count']))
    handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    handler.setLevel(config['LOGGING']['log_level'])  # Set handler log level dynamically

    # Set the root logger level dynamically based on config
    root_logger.setLevel(logging.DEBUG)

    # Add the handler to the logger
    root_logger.addHandler(handler)

    logger = logging.getLogger(dir_name)  

    # --------------------------------LOCAL LOGGING SETUP--------------------------------
    local_log_file_path = os.path.join(project_dir, dir_name, config[log_config]['log_file'])

    # Create a file handler for the PDFToHtmlConverter logs
    file_specific_handler = logging.handlers.RotatingFileHandler(local_log_file_path, maxBytes=int(config[log_config]['max_bytes']), backupCount=int(config[log_config]['backup_count']))
    file_specific_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    file_specific_handler.setLevel(config[log_config]['log_level']) 

    logger.addHandler(file_specific_handler)

    return logger


class EpicMetaTagComplianceValidator:
    def __init__(self, df):
        self.df = df
        self.unique_id_set = set()

        self.workbook = Workbook()
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet) #removes the default 'Sheet'

        # Create the Summary sheet as the first sheet
        self.summary_sheet = self.workbook.create_sheet(title="Summary", index=0)

        self.sheets = {
            "Title Check": self.workbook.create_sheet(title="Title Check"),
            "Unique Name Check": self.workbook.create_sheet(title="Unique Name Check"),
            "Keyword Check": self.workbook.create_sheet(title="Keyword Check"),
            "Language Check": self.workbook.create_sheet(title="Language Check"),
            "Filename Check": self.workbook.create_sheet(title="Filename Check"),
            "Diagnosis Code Check": self.workbook.create_sheet(title="Diagnosis Code Check"),
            "CPT Code Check": self.workbook.create_sheet(title="CPT Code Check")
        }
        
        # Initialize headers for each sheet
        for sheet_name, sheet in self.sheets.items():
            sheet.append(["File Directory Path", "Filename", "Issue Description", "Issue Status"])
            for col in sheet[1]:
                col.font = Font(bold=True, size=15)  # Make column headers bold
                sheet.column_dimensions[col.column_letter].width = 30  # Set column width
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping

    def setUp(self):
        self.logger = init_logging('LOGGING_TagValidator', 'EpicComplianceValidator')
        self.logger.info("RUNNING EPIC TAG VALIDATOR")
        self.errorCount = 0
        self.log_file = os.path.join(project_dir, config['LOGGING']['log_file'])
        self.errors = [] #Empty list to store the errors 
        #self.html_folder = os.path.join(configCustomer[f'{customer}']['local_folder_path'], 'Folder Storage/HTML')


    def update_sheet(self, sheet_name, filepath, filename, issue_description, issue_status):
        # Append the result for the file to the sheet
        self.sheets[sheet_name].append(
            [filepath, filename, issue_description, issue_status]
        )

        # Apply formatting to the sheet
        sheet = self.sheets[sheet_name]
        issue_status_column = 4  # Assuming "Issue Status" is the 4th column
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=issue_status_column, max_col=issue_status_column):
            for cell in row:
                if cell.value == "Issue Found":
                    cell.fill = red_fill
                elif cell.value == "No Issue":
                    cell.fill = green_fill

        # Adjust column width and enable text wrapping
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 30
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
    

    def check_for_illegal_chars(self, text, chars, tab_str):
        illegal_chars_found = []
        
        for char in chars:
            if char in text:
                illegal_chars_found.append(char)
        if tab_str in text:
            illegal_chars_found.append(tab_str)
        return illegal_chars_found

    def validate_title(self, total_issues, row_dict):
        filepath = row_dict[COLS['filepath']]
        filename = os.path.basename(filepath)
        errorlist = []
        sheet_name = "Title Check"

        # Logic Starts ############################################################################################
        title_content = row_dict[COLS['title']]
        chars_to_detect = "[],|^"
        tab = '<tab>'
        
        if not (type(title_content) == str and len(title_content) > 0):
            self.logger.error(f"Empty title tag in file {filename}\n")
            errorlist.append(f"Empty title tag in file {filename}")   
        else: 
            illegal_chars = self.check_for_illegal_chars(title_content, chars_to_detect, tab)           
            if illegal_chars:
                self.logger.error(f"Detected illegal character(s) in title tag in file {filename}: {illegal_chars}\n")
                errorlist.append(f"Detected illegal character(s) in title tag in file {filename}: {illegal_chars}")
            
            # Title tag should not include the language?
            # language_indicators = ['English', 'Spanish', 'ES']
            # for ind in language_indicators:
            #     if ind in title_content:
            #         self.logger.error(f"Detected language indicator in title tag in file {row_dict[COLS['filepath']]}: {ind}\n")
            #         errorlist.append(f"Detected language indicator in title tag in file {row_dict[COLS['filepath']]}: {ind}")
            #         break
        # Logic Ends ##############################################################################################

        # Determine issue status based on the errorlist
        if errorlist:
            issue_status = "Issue Found"
            issue_description = f"Title tag validation failed with {len(errorlist)} error(s): {', '.join(errorlist[:3])}"  # Show the first 3 errors
            total_issues[sheet_name] += 1
        else:
            issue_status = "No Issue"
            issue_description = "Title tag meets EPIC specifications"

        self.update_sheet(sheet_name, filepath, filename, issue_description, issue_status)


    def validate_unique_name(self, total_issues, row_dict):
        filepath = row_dict[COLS['filepath']]
        filename = os.path.basename(filepath)
        errorlist = []
        sheet_name = "Unique Name Check"

        # Logic Starts ############################################################################################
        unique_name = row_dict[COLS['unique_name']]
        if not (type(unique_name) == str and len(unique_name) > 0):
            self.logger.error(f"Empty unique name tag in file {filename}: {unique_name}\n")
            errorlist.append(f"Empty unique name tag in file {filename}: {unique_name}")
        elif len(unique_name) > 192 or not re.match(r'^[a-zA-Z0-9\-_.]+$', unique_name):
            self.logger.error(f"Invalid unique name tag in file {filename}: {unique_name}\n")
            errorlist.append(f"Invalid unique name tag in file {filename}: {unique_name}")
        else:
            if unique_name in self.unique_id_set:
                self.logger.error(f"Duplicate Unique ID in file {filename}: '{unique_name}'\n")
                errorlist.append(f"Duplicate Unique ID in file {filename}: '{unique_name}'")
            else:
                self.unique_id_set.add(unique_name)
        # Logic Ends ##############################################################################################
        
        # Determine issue status based on the errorlist
        if errorlist:
            issue_status = "Issue Found"
            issue_description = f"Unique name tag validation failed with {len(errorlist)} error(s): {', '.join(errorlist[:3])}"  # Show the first 3 errors
            total_issues[sheet_name] += 1
        else:
            issue_status = "No Issue"
            issue_description = "Unique ID is valid and unique"

        self.update_sheet(sheet_name, filepath, filename, issue_description, issue_status)


    def validate_keywords(self, total_issues, row_dict):
        filepath = row_dict[COLS['filepath']]
        filename = os.path.basename(filepath)
        errorlist = []
        sheet_name = "Keyword Check"

        # Logic Starts ############################################################################################
        keywords = row_dict[COLS['keyword']]
        if not (type(keywords) == str and len(keywords) > 0):
            self.logger.error(f"Empty Keyword Tag in file {filename}\n")
            errorlist.append(f"Empty Keyword Tag in file {filename}")
        else:
            keywords_pattern = re.compile(r'\b[A-Za-z ]+(?=,|\b)')
            if not keywords_pattern.match(keywords):
                self.logger.error(f"Invalid syntax for Keywords. Only comma-separated English words are allowed\n")
                errorlist.append(f"Invalid syntax for Keywords. Only comma-separated English words are allowed")
            else:
                keyword_list = keywords.split(',')
                for word in keyword_list:
                    if len(word) > 184:
                        self.logger.error(f"Keyword longer than 184 characters in file {filename}: {len(word)} characters\n")
                        errorlist.append(f"Keyword longer than 184 characters in file {filename}: {len(word)} characters")
                        break

                    # if detect(word) != 'en':
                    #     self.logger.error(f"Non-english keywords detected in file {filename}\n")
                    #     errorlist.append(f"Non-english keywords detected in file {filename}")
                    #     break
                
                if len(keywords) > 500:
                    self.logger.error(f"More than 500 keywords detected in file {filename}\n")
                    errorlist.append(f"More than 500 keywords detected in file {filename}")
        # Logic Ends ##############################################################################################
        
        # Determine issue status based on the errorlist
        if errorlist:
            issue_status = "Issue Found"
            issue_description = f"Keyword tag validation failed with {len(errorlist)} error(s): {', '.join(errorlist[:3])}"  # Show the first 3 errors
            total_issues[sheet_name] += 1
        else:
            issue_status = "No Issue"
            issue_description = "Keyword tag meets EPIC specifications"

        self.update_sheet(sheet_name, filepath, filename, issue_description, issue_status)


    def validate_language(self, total_issues, row_dict):
        filepath = row_dict[COLS['filepath']]
        filename = os.path.basename(filepath)
        errorlist = []
        sheet_name = "Language Check"

        # Logic Starts ############################################################################################
        language_content = row_dict[COLS['language']]
        if not (type(language_content) == str and len(language_content) > 0):
            self.logger.error(f'Empty Language Tag in file {filename}.')
            errorlist.append(f'Empty Language Tag in file {filename}.')
        else:
            # Check if 'ES' is in the filename
            has_es_in_filename = 'ES' in filename

            # Validate the conditions
            language_content = language_content.strip().lower()
            if not ((has_es_in_filename and language_content in ['spa', 'spanish']) or (not has_es_in_filename and language_content in ['eng', 'english'])):
                self.logger.error(f'Mismatch: Filename (\'ES\' in filename: {has_es_in_filename}) and Language meta tag (\'{language_content}\') do not align in file {filename}.')
                errorlist.append(f'Mismatch: Filename (\'ES\' in filename: {has_es_in_filename}) and Language meta tag (\'{language_content}\') do not align in file {filename}.')
                
            languages = language_content.strip().split()
            if len(languages) > 1:
                self.logger.error(f'Found multiple Language Tags in file {filename}.')
                errorlist.append(f'Found multiple Language Tags in file {filename}.')
        # Logic Ends ##############################################################################################
        
        # Determine issue status based on the errorlist
        if errorlist:
            issue_status = "Issue Found"
            issue_description = f"Language tag validation failed with {len(errorlist)} error(s): {', '.join(errorlist[:3])}"  # Show the first 3 errors
            total_issues[sheet_name] += 1
        else:
            issue_status = "No Issue"
            issue_description = "Language tag meets EPIC specifications"

        self.update_sheet(sheet_name, filepath, filename, issue_description, issue_status)


    def validate_filename(self, total_issues, row_dict):
        filepath = row_dict[COLS['filepath']]
        filename = os.path.basename(filepath)
        errorlist = []
        sheet_name = "Filename Check"

        # Logic Starts ############################################################################################
        # Special characters in title
        special_chars = ["'", '"', ":", '’', '&', '‘', '“', '`', ']', '[', '\\', '/', '^', '*', '{', '}']
        for sc in special_chars:
            if sc in filename:
                self.logger.error(f"Special character ({sc}) detected in file: {filename}")
                errorlist.append(f"Special character ({sc}) detected in file: {filename}")
                break
        # Logic Ends ##############################################################################################
        
        # Determine issue status based on the errorlist
        if errorlist:
            issue_status = "Issue Found"
            issue_description = f"Filename validation failed with {len(errorlist)} error(s): {', '.join(errorlist[:3])}"  # Show the first 3 errors
            total_issues[sheet_name] += 1
        else:
            issue_status = "No Issue"
            issue_description = "Filename meets EPIC specifications"

        self.update_sheet(sheet_name, filepath, filename, issue_description, issue_status)

    
    def validate_diagnosis_code(self, total_issues, row_dict):
        filepath = row_dict[COLS['filepath']]
        filename = os.path.basename(filepath)
        errorlist = []
        sheet_name = "Diagnosis Code Check"

        # Logic Starts ############################################################################################
        icd_codes = row_dict[COLS['diagnosis_code']]
        if not (type(icd_codes) == str and len(icd_codes) > 0):
            self.logger.error(f'Empty ICD Code Tag in file {filename}.')
            errorlist.append(f'Empty ICD Code Tag in file {filename}.')
        else:
            patterns = {'icd9_pattern': {"valid": [], "invalid": [], 'pattern': re.compile(r'^\d{3}(\.\d{1,2})?$')}, 
                        'icd10_cm_pattern': {"valid": [], "invalid": [], 'pattern': re.compile(r'^[A-Z]\d{2}(\.\w{1,4})?$')}, 
                        'icd10_ca_pattern': {"valid": [], "invalid": [], 'pattern': re.compile(r'^[A-Z][0-9]\.(\w{1,3})?$')}, 
                        'icd10_uk_pattern': {"valid": [], "invalid": [], 'pattern': re.compile(r'^[A-Z]{3}\.\d{3}$')}, 
                        'icd10_am_pattern': {"valid": [], "invalid": [], 'pattern': re.compile(r'^[A-Z][0-9]\.\w{1,4}([A-Z0-9])?$')}}
            
            # Split the input string by commas and strip whitespace
            codes = [code.strip() for code in icd_codes.split(",")]
            
            for key in patterns:
                pattern = patterns[key]['pattern']
                # Classify codes as valid or invalid
                patterns[key]['valid'] = [code for code in codes if pattern.match(code)]
                patterns[key]['invalid'] = [code for code in codes if not pattern.match(code)]

            check = False
            for key in patterns:
                if len(patterns[key]['invalid']) == 0:
                    check = True
                    break

            if not check:
                self.logger.error(f'Invalid ICD Codes detected in file {filename}.')
                errorlist.append(f'Invalid ICD Codes detected in file {filename}.')
        # Logic Ends ##############################################################################################
        
        # Determine issue status based on the errorlist
        if errorlist:
            issue_status = "Issue Found"
            issue_description = f"Diagnosis code validation failed with {len(errorlist)} error(s): {', '.join(errorlist[:3])}"  # Show the first 3 errors
            total_issues[sheet_name] += 1
        else:
            issue_status = "No Issue"
            issue_description = "Diagnosis codes meet EPIC specifications"

        self.update_sheet(sheet_name, filepath, filename, issue_description, issue_status)

    
    def validate_cpt_code(self, total_issues, row_dict):
        filepath = row_dict[COLS['filepath']]
        filename = os.path.basename(filepath)
        errorlist = []
        sheet_name = "CPT Code Check"

        # Logic Starts ############################################################################################
        cpt_codes = row_dict[COLS['cpt_code']]
        if not (type(cpt_codes) == str and len(cpt_codes) > 0):
            self.logger.error(f'Empty CPT Code Tag in file {filename}.')
            errorlist.append(f'Empty CPT Code Tag in file {filename}.')
        else:
            cpt_pattern = re.compile(r'^[A-Z0-9]{5}$')

            # Split the input string by commas and strip whitespace
            codes = [code.strip() for code in cpt_codes.split(",")]
            
            # Classify codes as valid or invalid
            valid_codes = [code for code in codes if cpt_pattern.match(code)]
            invalid_codes = [code for code in codes if not cpt_pattern.match(code)]

            if len(invalid_codes) > 0:
                self.logger.error(f'Invalid CPT Codes detected in file {filename}.')
                errorlist.append(f'Invalid CPT Codes detected in file {filename}.')
        # Logic Ends ##############################################################################################
        
        # Determine issue status based on the errorlist
        if errorlist:
            issue_status = "Issue Found"
            issue_description = f"CPT code validation failed with {len(errorlist)} error(s): {', '.join(errorlist[:3])}"  # Show the first 3 errors
            total_issues[sheet_name] += 1
        else:
            issue_status = "No Issue"
            issue_description = "CPT codes meet EPIC specifications"

        self.update_sheet(sheet_name, filepath, filename, issue_description, issue_status)
   

    def generate_summary(self):
        # Initialize the headers for the summary sheet
        self.summary_sheet.append(["File Directory Path", "Filename", "Issues Found"])
        for col in self.summary_sheet[1]:  # Style the headers
            col.font = Font(bold=True, size=14)  # Bold and larger font
            self.summary_sheet.column_dimensions[col.column_letter].width = 30  # Set column width

        # Create a dictionary to track issues for each file
        file_issues = {}

        # Iterate through all individual sheets
        for sheet_name, sheet in self.sheets.items():
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
                # Adjust unpacking based on the sheet
                if sheet_name == "Image Path Check":
                    file_path, filename, image_path, issue_description, issue_status = row
                else:
                    file_path, filename, issue_description, issue_status = row

                if issue_status == "Issue Found":
                    # Add the sheet name to the issues for this file
                    if (file_path, filename) not in file_issues:
                        file_issues[(file_path, filename)] = []
                    file_issues[(file_path, filename)].append(sheet_name)

        # Populate the summary sheet
        total_issues = 0  # Counter for total issues across all files
        processed_files = set()  # Keep track of processed files to avoid duplicates

        for (file_path, filename), issues_list in file_issues.items():
            if (file_path, filename) not in processed_files:
                issues = ", ".join(issues_list)  # Join sheet names with issues
                issue_status = "Issue Found"
                total_issues += 1  # Increment total issues
                # Append the row to the summary sheet
                self.summary_sheet.append([file_path, filename, issues])
                processed_files.add((file_path, filename))

        # Add entries for files without issues (only if they exist in the sheets)
        for sheet in self.sheets.values():
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                if len(row) == 5:  # Handle the Image Path Check sheet with 5 columns
                    file_path, filename, image_path, issue_description, issue_status = row
                else:  # Handle other sheets with 4 columns
                    file_path, filename, issue_description, issue_status = row

                if issue_status == "No Issue" and (file_path, filename) not in processed_files:
                    self.summary_sheet.append([file_path, filename, "No issues found"])
                    processed_files.add((file_path, filename))

        # Apply color formatting to the 'Issues Found' column
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        issue_column_index = 3  # The 'Issues Found' column is the 3rd column

        for row in self.summary_sheet.iter_rows(min_row=2, max_row=self.summary_sheet.max_row, min_col=issue_column_index, max_col=issue_column_index):
            for cell in row:
                if cell.value == "No issues found":
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

        # Add total issues row at the end of the summary sheet
        self.summary_sheet.append(["", "", f"Total Files with Issues: {total_issues}"])
        self.summary_sheet.append([""])  # Blank row for better readability

        # Enable text wrapping in the summary sheet
        for row in self.summary_sheet.iter_rows(min_row=2, max_row=self.summary_sheet.max_row, min_col=1, max_col=self.summary_sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)


    def validate_spreadsheet(self):
        total_issues = {sheet_name: 0 for sheet_name in self.sheets}
        self.setUp()

        for index, row in self.df.iterrows():
            row_num = index + 1
            row_dict = row.to_dict()
            self.validate_title(total_issues, row_dict)
            self.validate_unique_name(total_issues, row_dict)
            self.validate_keywords(total_issues, row_dict)
            self.validate_language(total_issues, row_dict)
            self.validate_filename(total_issues, row_dict)
            self.validate_diagnosis_code(total_issues, row_dict)
            self.validate_cpt_code(total_issues, row_dict)

        # Add totals row at the bottom of each sheet
        for sheet_name, sheet in self.sheets.items():
            total_files_with_issues = total_issues[sheet_name]
            sheet.append(["", "", "", f"Total Files with Issues: {total_files_with_issues}"])
            sheet.append([""])  # Blank row for better readability

        self.generate_summary()

        report_filename = f"{customer}_MetaTagSpreadsheetValidationReport.xlsx"

        # Save the workbook with the generated filename
        self.workbook.save(os.path.join(st.session_state['SPREADSHEET_DIR'], report_filename))
