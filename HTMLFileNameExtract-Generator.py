import os
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Specify the top-level directory
#top_level_dir = '/Users/manju/Mytonomy Inc Dropbox/Manju Komarlu/Written Education - Upwork Share/WE-Tester01/Carolina East Clinical References Root 10-8-24-Shipped/HTML'
#top_level_dir = '/Users/manju/Mytonomy Inc Dropbox/Manju Komarlu/Written Education - Upwork Share/WE-Tester01/AAH/Jan-2025-First100 Test-ENGLISH/HTML'
top_level_dir = '/Users/manju/Work/WrittenEducation Project/Customer Releases/AAH/HTML/Jan-20-2025-AAH_Test batch-113 Files/HTML/Clara Project/ENGLISH'

# Initialize an empty list to store file information
file_info = []

# Walk through the subfolders and capture file information
for root, dirs, files in os.walk(top_level_dir):
    # Skip the "Images" folder
    if 'Images' in dirs:
        dirs.remove('Images')

    for file in files:
        # Capture the file name and relative path
        file_name = file
        relative_path = os.path.relpath(root, top_level_dir)
        file_path = os.path.join(relative_path, file_name)

        # Initialize video URL, header text, and subdirectories to None
        video_url = None
        header_text = None
        subdirectories = None

        # Check if the file is an HTML file
        if file.endswith('.html'):
            # Extract the subdirectories
            subdirectories = '/'.join(relative_path.split('/')[:-1])

            # Open the HTML file and read its contents
            with open(os.path.join(root, file), 'r') as f:
                html_content = f.read()

            # Parse the HTML content using BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')

            # Find the div tags
            div_tags = soup.find_all('div')
            for div in div_tags:
                # Find the h1 or h3 tag within the div tag
                header_tag = div.find(['h1', 'h3', 'p'])
                if header_tag:
                    #header_text = header_tag.text.strip()
                    # Find the anchor element with the href pattern
                    a_tag = header_tag.find('a', href=lambda x: x and x.startswith('https://myto.us/'))
                    if a_tag:
                        # Extract the entire URL as the Video URL
                        video_url = a_tag['href']
                        break

        # Append the file information to the list
        file_info.append([subdirectories, file_path, video_url, header_text])

# Create a pandas DataFrame from the file information list
df = pd.DataFrame(file_info, columns=['Medical Category', 'File Path', 'Video URL', 'Module in PEC Portal after scanning the Mobile QR Code'])

# Generate a spreadsheet with the file information
#with pd.ExcelWriter('Carolina East Clinical References Root 10-8-24-Shipped.xlsx', engine='openpyxl') as writer:
with pd.ExcelWriter('Jan-2025-First100 Test-ENGLISH-QA.xlsx', engine='openpyxl') as writer:

    df.to_excel(writer, index=False)
    sheet = writer.sheets['Sheet1']
    for i in range(2, sheet.max_row + 1):
        if i % 2 == 0:
            sheet.cell(row=i, column=1).fill = PatternFill(start_color='C7C7C7', end_color='C7C7C7', fill_type='solid')
            sheet.cell(row=i, column=2).fill = PatternFill(start_color='C7C7C7', end_color='C7C7C7', fill_type='solid')
            sheet.cell(row=i, column=3).fill = PatternFill(start_color='C7C7C7', end_color='C7C7C7', fill_type='solid')
            sheet.cell(row=i, column=4).fill = PatternFill(start_color='C7C7C7', end_color='C7C7C7', fill_type='solid')
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
