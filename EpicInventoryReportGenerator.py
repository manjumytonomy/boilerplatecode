import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

spreadsheet_path = "/Users/vikram/Downloads/AAH_March_2025_EpicMetaData - Locked.xlsx"
package_path = "/Users/vikram/Mytonomy Inc Dropbox/Vikram Rao/WE-Tester01/AAH/March 2025 Delivery/AAH 303_03_16_2025-M-S"
report_path = "Report.xlsx"

html_dir = "/Users/vikram/Work/Mytonomy-Official/AAH/Fixing Master Spreadsheet/AAH A-B 186_02_27_2025 HTML/HTML"
pdf_dir = "/Users/vikram/Work/Mytonomy-Official/AAH/Fixing Master Spreadsheet/PDFs"

df = pd.read_excel(spreadsheet_path)
df['Filepath'] = df['Filepath'].astype(str)
df['Title'] = df['Title'].astype(str)
df = df.sort_values(by='Filepath').reset_index(drop=True)

pdf_files = []
for root, dirs, files in sorted(os.walk(pdf_dir)):
    dirs.sort()  # Sort directories alphabetically to ensure ordered traversal
    files.sort()  # Sort files alphabetically
    
    for file in files:
        if file.endswith(".pdf"):
            filepath = os.path.join(root, file)
            idx = len(pdf_dir)
            pdf_files.append(filepath[idx:])

html_files = []
for root, dirs, files in sorted(os.walk(html_dir)):
    dirs.sort()  # Sort directories alphabetically to ensure ordered traversal
    files.sort()  # Sort files alphabetically
    
    for file in files:
        if file.endswith(".html"):
            filepath = os.path.join(root, file)
            idx = len(html_dir)
            html_files.append(filepath[idx:])

letters = set()
for file in html_files:
    letters.add(file.split('/')[3])
letters = list(letters)

df['Letter'] = df['Filepath'].apply(lambda x: x.split('/')[3])
df = df[df['Letter'].isin(letters)]

report_data = []  # List to store the final data
html_filepaths = []

for index, row in df.iterrows():
    spreadsheet_filepath = row['Filepath']
    spreadsheet_title = row['Title']

    html_filepath = os.path.join(os.path.dirname(spreadsheet_filepath), spreadsheet_title + '.html')
    html_filepaths.append(html_filepath)

    # # Your custom logic for matching:
    # matched_pdf = next((pdf for pdf in pdf_files if spreadsheet_filepath in pdf), None)
    # matched_html = next((html for html in html_files if html_filepath in html), None)

    # Append the row as a list
    report_data.append([spreadsheet_filepath, pdf_files[index], spreadsheet_title, html_filepath, html_files[index]])

# mismatched_pdfs = [f for f in pdf_files if f not in df['Filepath'].to_list()]
# mismatched_htmls = [f for f in html_files if f not in html_filepaths]

# for file in mismatched_pdfs:
#     report_data.append([None, None, file, None])
# for file in mismatched_htmls:
#     report_data.append([None, None, None, file])

report_df = pd.DataFrame(report_data, columns=['Spreadsheet Filepath', 'Package PDF Filepath', 'Spreadsheet Title', 'HTML Filepath', 'Package HTML Filepath'])
report_df.to_excel(report_path)

# Load the workbook and select the active sheet
wb = load_workbook(report_path)
ws = wb.active

# Define the fill color for mismatched cells
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Get column indices for reference
spreadsheet_col = 2  # Column A (Spreadsheet Filepath)
package_pdf_col = 3   # Column B (Package PDF Filepath)
html_col = 5
package_html_col = 6

# Iterate through rows, skipping the header (start from row 2)
for row in range(2, ws.max_row + 1):
    spreadsheet_value = ws.cell(row=row, column=spreadsheet_col).value
    package_pdf_value = ws.cell(row=row, column=package_pdf_col).value
    html_value = ws.cell(row=row, column=html_col).value
    package_html_value = ws.cell(row=row, column=package_html_col).value

    if spreadsheet_value != package_pdf_value:
        ws.cell(row=row, column=package_pdf_col).fill = highlight_fill  # Highlight mismatched cell

    if html_value != package_html_value:
        ws.cell(row=row, column=package_html_col).fill = highlight_fill  # Highlight mismatched cell

# Save the modified Excel file
wb.save(report_path)
