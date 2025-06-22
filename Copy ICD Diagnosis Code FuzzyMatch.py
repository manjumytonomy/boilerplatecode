import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Load spreadsheets
spreadsheet1 = pd.read_excel('Mytonomy Merged A-Z HTML Package - May 2025 Release - MetaData Extract.xlsx')
spreadsheet2 = pd.read_excel('Written Asset Metatags - Inova_Rev.Aug2024.xlsx')

# Normalize titles
titles1 = spreadsheet1['Title'].astype(str).str.strip().str.lower()
titles2 = spreadsheet2['Title'].astype(str).str.strip().str.lower()

# TF-IDF vectorization
vectorizer = TfidfVectorizer().fit(titles1.tolist() + titles2.tolist())
tfidf1 = vectorizer.transform(titles1)
tfidf2 = vectorizer.transform(titles2)

# Cosine similarity matrix
cosine_sim_matrix = cosine_similarity(tfidf1, tfidf2)

# Find best matches
best_match_indices = cosine_sim_matrix.argmax(axis=1)
best_match_scores = cosine_sim_matrix.max(axis=1)

# Build matched columns
matched_titles = []
matched_icd_codes = []
for i, score in enumerate(best_match_scores):
    idx = best_match_indices[i]
    matched_titles.append(spreadsheet2.iloc[idx]['Title'])
    if score >= 0.20:
        matched_icd_codes.append(spreadsheet2.iloc[idx]['Diagnosis Code'])
    else:
        matched_icd_codes.append(None)

# Add new columns to spreadsheet1
spreadsheet1['matched_title'] = matched_titles
spreadsheet1['title_match_score'] = best_match_scores.round(4)
spreadsheet1['title_match_diagnosis_code'] = matched_icd_codes

# Reorder columns: include original 'Diagnosis Code' next to matched
core_columns = [
    'Title',
    'matched_title',
    'title_match_score',
    'title_match_diagnosis_code',
    'Diagnosis Code'
]
remaining_columns = [col for col in spreadsheet1.columns if col not in core_columns]
spreadsheet1 = spreadsheet1[core_columns + remaining_columns]

# Save to Excel first
output_file = 'Mytonomy_Title_Matched_Metadata.xlsx'
spreadsheet1.to_excel(output_file, index=False)

# Load workbook for conditional formatting
wb = load_workbook(output_file)
ws = wb.active

# Find the column indices (1-based for openpyxl)
header = [cell.value for cell in ws[1]]
idx_diagnosis_code = header.index('Diagnosis Code') + 1
idx_matched_code = header.index('title_match_diagnosis_code') + 1

# Apply conditional formatting: red font + yellow fill if codes mismatch
red_font = Font(color="9C0006")
yellow_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

for row in range(2, ws.max_row + 1):
    original = ws.cell(row=row, column=idx_diagnosis_code).value
    matched = ws.cell(row=row, column=idx_matched_code).value
    if original != matched and matched is not None:
        cell = ws.cell(row=row, column=idx_matched_code)
        cell.font = red_font
        cell.fill = yellow_fill

# Save the updated workbook
wb.save(output_file)